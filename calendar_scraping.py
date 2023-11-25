import time
import traceback

import openpyxl

from selenium import webdriver
from selenium.common import NoSuchElementException
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


# css 찾을때 까지 10초대기
def time_wait(num, code):
    try:
        wait = WebDriverWait(driver, num).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, code)))
    except:
        print(code, '태그를 찾지 못하였습니다.')
        driver.quit()
    return wait


# frame 변경 메소드
def switch_frame(frame):
    driver.switch_to.default_content()  # frame 초기화
    driver.switch_to.frame(frame)  # frame 변경


time_to_idx = {
    '월': 1, '화': 2, '수': 3, '목': 4, '금': 5, '토': 6, '일': 7,
    '월요일': 1, '화요일': 2, '수요일': 3, '목요일': 4, '금요일': 5, '토요일': 6, '일요일': 7
}

try:
    workbook = openpyxl.load_workbook('/Users/jihoon/Study/Django/dosuri_hospitals.xlsx')
    sheet = workbook.active

    start_row = 54
    end_row = 7200
    driver = webdriver.Firefox()

    for row_idx, row in enumerate(sheet.iter_rows(values_only=True, min_row=start_row, max_row=end_row, max_col=3)):
        if row[2] is not None:
            print(row[0], 'pass')
            continue
        try:
            location = row[1]
            hospital = row[0]

            print(hospital)

            url = "https://map.naver.com/p?c=15.00,0,0,0,dh"
            driver.get(url)
            time.sleep(2)

            # css를 찾을때 까지 10초 대기
            time_wait(3, 'div.input_box > input.input_search')

            # (1) 검색창 찾기
            search = driver.find_element(By.CSS_SELECTOR, 'div.input_box > input.input_search')
            search.send_keys(location)  # 검색어 입력
            search.send_keys(Keys.ENTER)  # 엔터버튼 누르기

            time.sleep(2)

            search.clear()
            search.send_keys(hospital)  # 검색어 입력
            search.send_keys(Keys.ENTER)  # 엔터버튼 누르기

            time.sleep(5)

            try:
                print('entryIframe')
                # (2) frame 변경
                switch_frame('entryIframe')
                time.sleep(2)
                # 시간 열기
                target = driver.find_element(By.XPATH,
                                             '/html/body/div[3]/div/div/div/div[5]/div/div[1]/div/div/div[3]/div/a/div[1]/div/span').click()

                # 검색 결과가 하나인 경우
                print("검색 결과 한 개")

                time.sleep(2)

                time_loc_index = 3
                # for i in range(5):
                #     try:
                #         driver.find_element(By.XPATH,
                #                             f'/html/body/div[3]/div/div/div/div[5]/div/div[1]/div/div/div[{i}]/div/a/div/div/span').click()
                #         time_loc_index = i
                #         break
                #     except:
                #         pass
                #
                # time.sleep(1)
                time_dict = {}
                for i in range(2, 9):
                    day_data = driver.find_element(By.XPATH,
                                                   f'/html/body/div[3]/div/div/div/div[5]/div/div[1]/div/div/div[{time_loc_index}]/div/a/div[{i}]/div/span/span')
                    try:
                        time_data = driver.find_element(By.XPATH,
                                                        f'/html/body/div[3]/div/div/div/div[5]/div/div[1]/div/div/div[{time_loc_index}]/div/a/div[{i}]/div/span/div')

                    except Exception:
                        time_data = driver.find_element(By.XPATH,
                                                        f'/html/body/div[3]/div/div/div/div[5]/div/div[1]/div/div/div[{time_loc_index}]/div/a/div[{i}]/div/span/div')
                    print(day_data.text)
                    print(time_data.text)
                    time_dict[time_to_idx[day_data.text]] = time_data.text.split('\n')[0]
            except:
                # traceback.print_exc()
                print('searchIframe')
                switch_frame('searchIframe')
                # 첫번째 검색결과 클릭
                try:
                    driver.find_element(By.XPATH,
                                        '/html/body/div[3]/div/div[3]/div[1]/ul/li[1]/div[1]/a[1]').click()
                except:
                    driver.find_element(By.XPATH,
                                        '/html/body/div[3]/div/div[3]/div[1]/ul/li[1]/div[1]/div/a[1]').click()

                # # 검색 결과가 여러개인 경우
                # # (2) frame 변경
                # switch_frame('entryIframe')
                #
                # time_loc_index = 2
                # for i in range(5):
                #     try:
                #         driver.find_element(By.XPATH,
                #                             f'/html/body/div[3]/div/div/div/div[5]/div/div[1]/div/div/div[{i}]/div/a/div/div/span').click()
                #         time_loc_index = i
                #         break
                #     except:
                #         pass


                time.sleep(2)
                switch_frame('entryIframe')

                time_dict = {}
                try:
                    time_loc_index = 2
                    # 시간 열기
                    driver.find_element(By.XPATH,
                                        f'/html/body/div[3]/div/div/div/div[5]/div/div[1]/div/div/div[{time_loc_index}]/div/a/div/div/span').click()
                    for i in range(2, 9):
                        day_data = driver.find_element(By.XPATH,
                                                       f'/html/body/div[3]/div/div/div/div[5]/div/div[1]/div/div/div[{time_loc_index}]/div/a/div[{i}]/div/span/span')
                        time_data = driver.find_element(By.XPATH,
                                                        f'/html/body/div[3]/div/div/div/div[5]/div/div[1]/div/div/div[{time_loc_index}]/div/a/div[{i}]/div/span/div')
                        time_dict[time_to_idx[day_data.text]] = time_data.text.split('\n')[0]

                except:
                    time_loc_index = 3
                    # 시간 열기
                    driver.find_element(By.XPATH,
                                        # '/html/body/div[3]/div/div/div/div[5]/div/div[1]/div/div/div[2]/div/a/div/div/span'
                                        # '/html/body/div[3]/div/div/div/div[5]/div/div[1]/div/div/div[2]/div/a/div/div/span'
                                        f'/html/body/div[3]/div/div/div/div[5]/div/div[1]/div/div/div[{time_loc_index}]/div/a/div/div/span').click()
                    for i in range(2, 9):
                        day_data = driver.find_element(By.XPATH,
                                                       f'/html/body/div[3]/div/div/div/div[5]/div/div[1]/div/div/div[{time_loc_index}]/div/a/div[{i}]/div/span/span')
                        time_data = driver.find_element(By.XPATH,
                                                        f'/html/body/div[3]/div/div/div/div[5]/div/div[1]/div/div/div[{time_loc_index}]/div/a/div[{i}]/div/span/div')
                        time_dict[time_to_idx[day_data.text]] = time_data.text.split('\n')[0]

            for key, value in time_dict.items():
                sheet.cell(row=row_idx + start_row, column=key + 2, value=value)
            time.sleep(5)
            workbook.save('/Users/jihoon/Study/Django/dosuri_hospitals.xlsx')
        except:
            traceback.print_exc()
            pass
        # finally:
        # workbook.save('/Users/jihoon/Study/Django/dosuri_hospitals.xlsx')
        # driver.quit()

except Exception:
    pass
    traceback.print_exc()
    # driver.quit()
