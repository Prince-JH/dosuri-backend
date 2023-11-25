import traceback

import openpyxl

from django.core.management.base import BaseCommand
from dosuri.hospital.models import Hospital, HospitalCalendar
import re

# 정규식 패턴 작성
# 패턴: HH:MM - HH:MM 형식 (HH는 00-23, MM은 00-59)
time_format_pattern = r'^([0-1]?[0-9]|2[0-3]):([0-5]?[0-9]) - ([0-1]?[0-9]|2[0-3]):([0-5]?[0-9])$'

# 예시 문자열을 이용한 정규식 테스트
example_strings = ["09:00 - 18:00", "23:59 - 00:00", "09:00-18:00", "9:00 - 18:00", "09:60 - 18:00"]

# 각 문자열에 대해 정규식 확인
matches = {string: bool(re.match(time_format_pattern, string)) for string in example_strings}
matches


class Command(BaseCommand):
    def validate_format(self, time: str):
        if not time:
            return None
        # 정규식 패턴 작성
        # 패턴: HH:MM - HH:MM 형식 (HH는 00-23, MM은 00-59)
        time_format_pattern = r'^([0-1]?[0-9]|2[0-3]):([0-5]?[0-9]) - ([0-1]?[0-9]|2[0-3]):([0-5]?[0-9])$'
        if re.match(time_format_pattern, time):
            return time.replace('-', '~')
        else:
            return None

    def handle(self, *args, **options):
        workbook = openpyxl.load_workbook('/Users/jihoon/Study/Django/dosuri_hospitals.xlsx')
        sheet = workbook['list']
        y = 0
        n = 0
        start_row = 3

        for row_idx, row in enumerate(
                sheet.iter_rows(values_only=True, min_row=start_row, max_row=7600, max_col=11)):
            if not row[2]:
                continue
            if row[10] == 'SUCCESS':
                continue

            code = row[9]
            monday = self.validate_format(row[2])
            tuesday = self.validate_format(row[3])
            wednesday = self.validate_format(row[4])
            thursday = self.validate_format(row[5])
            friday = self.validate_format(row[6])
            saturday = self.validate_format(row[7])
            sunday = self.validate_format(row[8])
            # print(monday, tuesday, wednesday, thursday, friday, saturday, sunday)

            # if row[2]:
            #     y += 1
            # else:
            #     n += 1
            try:
                print(code)
                hospital = Hospital.objects.get(code=code)

                qs = HospitalCalendar.objects.filter(hospital=hospital)
                if not qs.exists():
                    HospitalCalendar.objects.create(
                        hospital=hospital,
                        monday=monday,
                        tuesday=tuesday,
                        wednesday=wednesday,
                        thursday=thursday,
                        friday=friday,
                        saturday=saturday,
                        sunday=sunday
                    )
                sheet.cell(row=row_idx + start_row, column=11, value='SUCCESS')

            except:
                traceback.print_exc()
                sheet.cell(row=row_idx + start_row, column=11, value='FAIL')
                break
            # break
        workbook.save('/Users/jihoon/Study/Django/dosuri_hospitals.xlsx')
        # print(y, n)
